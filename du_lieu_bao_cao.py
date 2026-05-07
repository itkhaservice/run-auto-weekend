import os
import sys
import subprocess
import logging
import pandas as pd
from playwright.sync_api import sync_playwright, Page
import pytest
from openpyxl import load_workbook
from datetime import datetime
from tabulate import tabulate

# Cấu hình BASE_DIR
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

@pytest.fixture(scope="session")
def browser():
    with sync_playwright() as p:
        # Chạy headless=True trên CI, False trên local để debug
        is_ci = os.environ.get("CI") == "true"
        browser = p.chromium.launch(
            headless=is_ci,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-blink-features=AutomationControlled"]
        )
        yield browser
        browser.close()

@pytest.fixture
def page(browser):
    context = browser.new_context(viewport={'width': 1920, 'height': 1080})
    page = context.new_page()
    yield page
    context.close()

def login(page: Page):
    if "login" in page.url or "qlvh.khaservice.com.vn" not in page.url:
        page.goto("https://qlvh.khaservice.com.vn/login")
        page.locator("input[name='email']").fill("admin@khaservice.com.vn")
        page.locator("input[name='password']").fill("Kha@@123")
        page.locator("button[type='submit']").click()
        page.wait_for_timeout(2000)

# --- 1. LẤY OVERVIEW (Cột B, C, D, E) ---
def test_lay_thong_tin_du_an(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao", header=None)
    project_list = project_df.iloc[1:, 0].tolist()
    wb = load_workbook(excel_path)
    ws = wb["BaoCao"]
    login(page)
    
    for idx, project_val in enumerate(project_list, start=2):
        try:
            print(f"[{idx}] Đang lấy Overview: {project_val}")
            page.locator("#combo-box-demo").click()
            page.locator("#combo-box-demo").fill(str(project_val))
            page.locator("#combo-box-demo-option-0").click()
            page.locator("a[href='/statistics/overview']").click()
            page.wait_for_load_state("networkidle")
            page.wait_for_timeout(1000)
            
            ws[f"B{idx}"] = page.locator('//*[@id="root"]/div[2]/main/div/div/div/div[2]/div/div[1]/p[1]').inner_text()
            ws[f"C{idx}"] = page.locator('//*[@id="root"]/div[2]/main/div/div/div/div[3]/div/div[1]/p[1]').inner_text()
            ws[f"D{idx}"] = page.locator('//*[@id="root"]/div[2]/main/div/div/div/div[5]/div/div[1]/p[1]').inner_text()
            ws[f"E{idx}"] = page.locator('//*[@id="root"]/div[2]/main/div/div/div/div[6]/div/div[1]/p[1]').inner_text()
        except Exception as e:
            print(f"Lỗi Overview {project_val}: {e}")
            
    wb.save(excel_path)

# --- 2. LẤY SỐ LƯỢNG BÀI VIẾT (Cột F, G) ---
def test_lay_so_luong_bai_viet(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao", header=None)
    project_list = project_df.iloc[1:, 0].tolist()
    wb = load_workbook(excel_path)
    ws = wb["BaoCao"]
    login(page)
    
    for idx, project_val in enumerate(project_list, start=2):
        try:
            print(f"[{idx}] Đang lấy Posts: {project_val}")
            page.locator("#combo-box-demo").click()
            page.locator("#combo-box-demo").fill(str(project_val))
            page.locator("#combo-box-demo-option-0").click()
            page.wait_for_timeout(1000)

            # --- TIN TỨC ---
            page.goto("https://qlvh.khaservice.com.vn/posts/news")
            page.wait_for_load_state("networkidle")
            
            # Chọn 1000 dòng
            try:
                page.locator("xpath=//*[@id='root']/div[2]/main/div/div/div[3]/div/div[2]/button").click()    
                page.locator("xpath=//*[@id='menu-apartment-list-style1']/div[3]/ul/li[6]").click()
                print("   -> Đã chọn hiển thị 1000 dòng tin tức...")
                page.wait_for_timeout(5000) # Chờ 5s để bảng load lại
            except: 
                print("   -> Không chọn được dropdown 1000 dòng")

            count_news = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr').count()
            ws[f"F{idx}"] = count_news
            print(f"   -> Tin tức: {count_news}")

            # --- THÔNG BÁO ---
            page.goto("https://qlvh.khaservice.com.vn/posts/notification")
            page.wait_for_load_state("networkidle")

            # Chọn 1000 dòng
            try:
                page.locator("xpath=//*[@id='root']/div[2]/main/div/div/div[3]/div/div[2]/button").click()    
                page.locator("xpath=//*[@id='menu-apartment-list-style1']/div[3]/ul/li[6]").click()
                print("   -> Đã chọn hiển thị 1000 dòng tin tức...")
                page.wait_for_timeout(5000) # Chờ 5s để bảng load lại
            except:
                print("   -> Không chọn được dropdown 1000 dòng")

            count_notif = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr').count()
            ws[f"G{idx}"] = count_notif
            print(f"   -> Thông báo: {count_notif}")

        except Exception as e:
            print(f"Lỗi Posts {project_val}: {e}")
            
    wb.save(excel_path)

# --- 3. LẤY NGÀY BÀI VIẾT CUỐI (Cột H) ---
def test_lay_thong_tin_bai_viet_ngay_cuoi(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao", header=None)
    project_list = project_df.iloc[1:, 0].tolist()
    wb = load_workbook(excel_path)
    ws = wb["BaoCao"]
    login(page)
    
    for idx, project_val in enumerate(project_list, start=2):
        try:
            print(f"[{idx}] Đang lấy Ngày cuối: {project_val}")
            page.locator("#combo-box-demo").click()
            page.locator("#combo-box-demo").fill(str(project_val))
            page.locator("#combo-box-demo-option-0").click()
            
            dates = []
            
            # Kiểm tra Thông báo (Lấy dòng đầu tiên)
            page.goto("https://qlvh.khaservice.com.vn/posts/notification")
            page.wait_for_load_state("networkidle")
            loc = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr[1]/td[8]/div')
            if loc.is_visible():
                text = loc.inner_text().strip()
                # Xử lý chuỗi ngày (ví dụ: "12/01/2026 14:00" -> lấy "12/01/2026")
                try:
                    date_str = text.split()[0]
                    dates.append(datetime.strptime(date_str, '%d/%m/%Y'))
                except: pass
            
            # Kiểm tra Tin tức (Lấy dòng đầu tiên)
            page.goto("https://qlvh.khaservice.com.vn/posts/news")
            page.wait_for_load_state("networkidle")
            loc = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr[1]/td[8]/div')
            if loc.is_visible():
                text = loc.inner_text().strip()
                try:
                    date_str = text.split()[0]
                    dates.append(datetime.strptime(date_str, '%d/%m/%Y'))
                except: pass
            
            if dates:
                max_date = max(dates).strftime('%d/%m/%Y')
                ws[f"H{idx}"] = max_date
                print(f"   -> Ngày mới nhất: {max_date}")
            else:
                ws[f"H{idx}"] = "N/A"
                
        except Exception as e:
            print(f"Lỗi Date {project_val}: {e}")
            
    wb.save(excel_path)

# --- 4. LẤY BÁO PHÍ MỚI NHẤT (Cột I) ---
def test_lay_thong_tin_bao_phi_moi_nhat(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao", header=None)
    project_list = project_df.iloc[1:, 0].tolist()
    wb = load_workbook(excel_path)
    ws = wb["BaoCao"]
    login(page)
    
    for idx, project_val in enumerate(project_list, start=2):
        try:
            print(f"[{idx}] Đang lấy Fee Report: {project_val}")
            page.locator("#combo-box-demo").click()
            page.locator("#combo-box-demo").fill(str(project_val))
            page.locator("#combo-box-demo-option-0").click()
            page.wait_for_timeout(1000)

            page.goto("https://qlvh.khaservice.com.vn/fee-reports")
            page.wait_for_load_state("networkidle")
            
            loc = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr[1]/td[5]/div')
            if loc.is_visible():
                text = loc.text_content().strip()
                ws[f"I{idx}"] = text
                print(f"   -> Phí mới nhất: {text}")
        except Exception as e:
            print(f"Lỗi Fee {project_val}: {e}")

    wb.save(excel_path)

# --- XUẤT BÁO CÁO RA GITHUB ---
def test_z_summary_report():
    excel_path = os.path.join(BASE_DIR, "data.xlsx")
    if not os.path.exists(excel_path): return
    
    # Đọc lại file Excel để lấy dữ liệu mới nhất
    df = pd.read_excel(excel_path, sheet_name="BaoCao")
    
    # --- CHUẨN HÓA DỮ LIỆU CHO JSON (Khớp với reportService.js) ---
    # Đổi tên cột từ Excel sang tên mà Website đang đọc
    # Cấu trúc cột thực tế: ['Dự án', 'Tổng căn hộ', 'Tổng cư dân sử dụng APP', 'Tổng số căn hộ sử dụng APP', 'Tổng số cư dân', 'Tin tức', 'Thông báo', 'Ngày mới nhất']
    
    mapping = {
        "Dự án": "Dự án",
        "Tổng căn hộ": "Tổng căn hộ",
        "Tổng cư dân sử dụng APP": "Tổng cư dân sử dụng APP",
        "Tổng số căn hộ sử dụng APP": "Tổng số căn hộ sử dụng APP",
        "Tin tức": "Tin tức",
        "Thông báo": "Thông báo",
        "Ngày mới nhất": "Ngày mới nhất"
    }
    
    # Chỉ rename những cột tồn tại
    df_json = df.rename(columns=mapping)
    
    # Nếu thiếu cột 'Báo phí', thêm vào với giá trị N/A (Vì trang web cần cột này)
    if "Báo phí" not in df_json.columns:
        # Thử tìm cột cuối cùng nếu nó là cột báo phí nhưng tên khác
        if len(df.columns) > 8:
            df_json["Báo phí"] = df.iloc[:, 8]
        else:
            df_json["Báo phí"] = "N/A"

    # Chuyển đổi dữ liệu thành bảng Markdown cho GitHub Summary
    table = tabulate(df, headers='keys', tablefmt='github', showindex=False)
    
    # --- XUẤT RA JSON ---
    json_path = os.path.join(BASE_DIR, "report.json")
    # Xuất toàn bộ dữ liệu (Website sẽ tự slice(1) để bỏ dòng đầu nếu cần)
    df_json.to_json(json_path, orient='records', force_ascii=False, indent=4)
    print(f"   -> Đã xuất báo cáo ra file: {json_path}")

    output = f"## 📊 Báo Cáo Tổng Hợp Dữ Liệu\n"
    output += f"*Thời gian cập nhật: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}*\n\n"
    output += table
    output += "\n\n---\n💡 **Hướng dẫn:** Bôi đen bảng dữ liệu ở trên, nhấn `Ctrl+C`, sau đó mở Excel và nhấn `Ctrl+V` để dán."

    if 'GITHUB_STEP_SUMMARY' in os.environ:
        with open(os.environ['GITHUB_STEP_SUMMARY'], 'a', encoding='utf-8') as f:
            f.write(output)
    else:
        print("\n" + output + "\n")
